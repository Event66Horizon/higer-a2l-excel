import os
import tkinter as tk
from tkinter import ttk
import pandas as pd
from openpyxl import load_workbook


# ---------- 1. 读取有效 BOM ----------
def read_valid_bom_rows(xlsx_file: str) -> pd.DataFrame:
    df = pd.read_excel(xlsx_file, sheet_name=0, dtype=str, header=None)

    header_idx = None
    for idx, row in df.iterrows():
        if str(row.iloc[0]).strip() == '序号':
            header_idx = idx
            break
    if header_idx is None:
        raise ValueError('未找到“序号”表头')

    data_df = df.iloc[header_idx + 1:].reset_index(drop=True)
    data_df.columns = df.iloc[header_idx].fillna('').astype(str).str.strip()

    for idx, row in data_df.iterrows():
        if row.dropna().astype(str).str.strip().eq('').all():
            data_df = data_df.iloc[:idx]
            break

    valid = data_df[data_df['序号'].notna() &
                    (data_df['序号'].astype(str).str.strip() != '')]
    return valid


# ---------- 2. 把 BOM 写进目标文件 ----------
def overwrite_target_with_bom(target_file: str, bom_rows: pd.DataFrame):
    wb = load_workbook(target_file)
    ws = wb.active

    # 清空第 3 行及以下
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    row_num = 3
    for _, r in bom_rows.iterrows():
        erp_code = str(r['存货编码']).strip()

        if erp_code.startswith('Y10000'):
            try:
                suffix = int(erp_code[-2:])
                offline_ver = f'V0.{suffix + 1}'
            except ValueError:
                offline_ver = 'V0.1'
        else:
            offline_ver = 'V0.1'

        ws.cell(row=row_num, column=1,  value=str(r.iloc[-1]).strip())   # 规格型号
        ws.cell(row=row_num, column=2,  value='半成品')                  # 大类
        ws.cell(row=row_num, column=3,  value='内部线束')                # 中类
        ws.cell(row=row_num, column=4,  value=str(r['零件名称']).strip())  # 小类
        ws.cell(row=row_num, column=5,  value='正式')                    # 零件状态
        ws.cell(row=row_num, column=6,  value=erp_code)                  # ERP编码
        # 第7列不动
        ws.cell(row=row_num, column=8,  value=str(r['零件名称']).strip())  # 零件名称
        ws.cell(row=row_num, column=9,  value='A.1')                     # 版本
        ws.cell(row=row_num, column=10, value=offline_ver)               # 线下版本
        ws.cell(row=row_num, column=11, value=str(r['单位']).strip())     # 计量单位
        ws.cell(row=row_num, column=12, value='自制')                    # 制造方式
        row_num += 1

    wb.save(target_file)
    return len(bom_rows)


# ---------- 3. 扫描当前目录所有 xlsx ----------
def scan_xlsx() -> list[str]:
    return [f for f in os.listdir('.') if f.lower().endswith('.xlsx')]


# ---------- 4. GUI ----------
def gui_interface():
    xlsx_list = scan_xlsx()
    if not xlsx_list:
        tk.messagebox.showerror('错误', '当前目录没有任何 .xlsx 文件！')
        return

    root = tk.Tk()
    root.title("部件导入工具")
    root.geometry("500x250")
    root.resizable(False, False)

    tk.Label(root, text="待处理文件（BOM 源）：").pack(pady=5)
    src_combo = ttk.Combobox(root, values=xlsx_list, width=50, state='readonly')
    src_combo.current(0)
    src_combo.pack()

    tk.Label(root, text="要被覆写的文件：").pack(pady=5)
    dst_combo = ttk.Combobox(root, values=xlsx_list, width=50, state='readonly')
    dst_combo.current(0 if len(xlsx_list) == 1 else 1)  # 默认第二个
    dst_combo.pack()

    info_lbl = tk.Label(root, text="", fg="green")
    info_lbl.pack(pady=10)

    def run():
        src = src_combo.get()
        dst = dst_combo.get()
        if not src or not dst:
            info_lbl.config(text="两个文件都必须选择！", fg="red")
            return
        if src == dst:
            info_lbl.config(text="源文件与目标文件不能相同！", fg="red")
            return
        try:
            bom = read_valid_bom_rows(src)
            cnt = overwrite_target_with_bom(dst, bom)
            info_lbl.config(text=f"成功覆写 {cnt} 条记录 → {dst}", fg="green")
        except Exception as e:
            info_lbl.config(text=f"失败：{e}", fg="red")

    tk.Button(root, text="开始覆写", command=run, width=20, bg="green", fg="white").pack(pady=10)
    root.mainloop()


if __name__ == "__main__":
    gui_interface()