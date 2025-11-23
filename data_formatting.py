import pandas as pd
import os
import tkinter as tk
from openpyxl import load_workbook
import sys, os
import sys, os
# 打包后 exe 旁路径；没打包则是脚本所在路径
BASE_DIR = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) \
           else os.path.dirname(os.path.abspath(__file__))
# 如果是打包运行，_MEIPASS 存在；没打包就用 __file__ 所在目录
template_file = os.path.join(BASE_DIR, "Bom导入模板0.3.xlsx")
output_file   = os.path.join(BASE_DIR, "processed_output.xlsx")
sample_file   = os.path.join(BASE_DIR, "sample.xlsx")

# ---------- 公共数据处理 ----------
def build_result_df():
    """把 sample.xlsx 处理成最终结果 DataFrame"""
    df = pd.read_excel("sample.xlsx", engine='openpyxl')[["层次", "规格", "数量"]]
    df['层次'] = pd.to_numeric(df['层次'], errors='coerce')
    df = df.sort_values(by='层次').reset_index(drop=True)

    new_rows = []
    main_levels = df[df['层次'] == df['层次'].astype(int)].copy()
    for _, main_row in main_levels.iterrows():
        main_level = int(main_row['层次'])
        new_rows.append({'层次': float(main_level),
                         '规格': main_row['规格'],
                         '数量': main_row['数量']})
        sub_df = df[(df['层次'] > main_level) & (df['层次'] < main_level + 1)].copy()

        def process_spec(spec):
            return '120Ω，1/4W' if '120Ω' in str(spec) else spec
        sub_df['规格'] = sub_df['规格'].apply(process_spec)
        sub_group = sub_df.groupby('规格', as_index=False)['数量'].sum()

        for idx, (_, row) in enumerate(sub_group.iterrows(), start=1):
            new_rows.append({'层次': float(f"{main_level}.{idx}"),
                             '规格': row['规格'],
                             '数量': row['数量']})
    return pd.DataFrame(new_rows)


# ---------- 写回模板 ----------
def write_to_template():
    try:
        template_file = os.path.join(BASE_DIR, "Bom导入模板0.3.xlsx")
        if not os.path.exists(template_file):
            return f"未找到模板文件：{template_file}"

        result_df = build_result_df()
        wb = load_workbook(template_file)
        ws = wb.active

        # 清空第 3 行及以下
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.value = None

        # 写入新数据
        for r_idx, row in enumerate(result_df.itertuples(index=False), start=3):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        wb.save(template_file)
        return f"已写回模板：{template_file}"
    except Exception as e:
        return f"写回模板失败：{e}"


# ---------- 输出到 processed_output.xlsx ----------
def write_to_processed():
    try:
        output_file = "processed_output.xlsx"
        build_result_df().to_excel(output_file, index=False, engine='openpyxl')
        return f"已输出到：{output_file}"
    except Exception as e:
        return f"输出失败：{e}"


# ---------- GUI ----------
def run_gui():
    def check_file():
        ok = os.path.exists("sample.xlsx")
        status = "已找到 sample.xlsx" if ok else "未找到 sample.xlsx"
        color = "green" if ok else "red"
        label_status.config(text=status, fg=color)
        btn_template.config(state="normal" if ok else "disabled")
        btn_processed.config(state="normal" if ok else "disabled")
        return ok

    def on_template():
        msg = write_to_template()
        show_msg(msg)

    def on_processed():
        msg = write_to_processed()
        show_msg(msg)

    def show_msg(msg):
        text_result.config(state='normal')
        text_result.delete(1.0, tk.END)
        text_result.insert(tk.END, msg)
        text_result.config(state='disabled')

    root = tk.Tk()
    root.title("BOM格式化工具")
    root.geometry("450x280")

    tk.Label(root, text="请确保当前目录下有：sample.xlsx").pack(pady=2)
    label_status = tk.Label(root, text="", font=("Arial", 12))
    label_status.pack(pady=5)
    tk.Button(root, text="刷新检测", command=check_file).pack(pady=5)

    btn_frame = tk.Frame(root)
    btn_frame.pack(pady=10)
    btn_template = tk.Button(btn_frame, text="写回模板\n(Bom导入模板0.3.xlsx)", width=22, command=on_template)
    btn_template.pack(side="left", padx=5)
    btn_processed = tk.Button(btn_frame, text="输出到新文件\n(processed_output.xlsx)", width=22, command=on_processed)
    btn_processed.pack(side="left", padx=5)

    text_result = tk.Text(root, height=6, width=60, state='disabled')
    text_result.pack(pady=10)

    check_file()
    root.mainloop()


if __name__ == "__main__":
    run_gui()